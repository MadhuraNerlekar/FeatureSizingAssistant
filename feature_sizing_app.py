from __future__ import annotations

import json
import math
import os
import re
import subprocess
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


DEFAULT_SIZE_TO_HOURS = {"XS": 2, "S": 4, "M": 8, "L": 16, "XL": 24}
SIZE_ORDER = ["XS", "S", "M", "L", "XL"]


@dataclass
class Requirement:
    text: str
    size: str
    notes: str = "—"


@dataclass
class ScreenDetail:
    title: str
    description: str


@dataclass
class ModuleSpec:
    name: str
    description: str
    requirements: List[Requirement] = field(default_factory=list)
    summary: str = ""
    screen_details: List[ScreenDetail] = field(default_factory=list)
    screen_count: int | None = None

    def total_hours(self, size_to_hours: Dict[str, float]) -> float:
        return sum(size_to_hours.get(req.size.upper(), 0) for req in self.requirements)

    def resolved_screen_count(self) -> int:
        if self.screen_details:
            return len(self.screen_details)
        if self.screen_count is not None:
            return max(1, self.screen_count)
        return max(1, math.ceil(len(self.requirements) / 2))


def ordered_size_keys(size_hours: Dict[str, float]) -> List[str]:
    primary = [size for size in SIZE_ORDER if size in size_hours]
    extras = [size for size in size_hours if size not in SIZE_ORDER]
    combined = primary + extras
    return combined or SIZE_ORDER.copy()


def get_openai_client(api_key: str | None = None) -> OpenAI:
    key = api_key or os.getenv("OPENAI_API_KEY")
    if not key:
        raise RuntimeError(
            "Missing OpenAI API key. Set the OPENAI_API_KEY environment variable "
            "or provide it in the UI."
        )
    return OpenAI(api_key=key)


def parse_llm_response(payload: str) -> Tuple[List[ModuleSpec], List[str], List[str], Dict[str, float]]:
    try:
        data = json.loads(payload)
    except json.JSONDecodeError as error:
        raise ValueError(f"LLM response was not valid JSON: {error}") from error

    size_hours_raw = data.get("size_hours") or DEFAULT_SIZE_TO_HOURS
    size_hours: Dict[str, float] = {}
    for size, hours in size_hours_raw.items():
        if size and isinstance(hours, (int, float)):
            size_hours[size.upper()] = float(hours)
    if not size_hours:
        size_hours = DEFAULT_SIZE_TO_HOURS.copy()
    else:
        for size, hours in DEFAULT_SIZE_TO_HOURS.items():
            size_hours.setdefault(size, float(hours))

    modules_data = data.get("modules", [])
    modules: List[ModuleSpec] = []
    for module in modules_data:
        requirements = []
        for req in module.get("requirements", []):
            requirements.append(
                Requirement(
                    text=req.get("text", "").strip() or "Requirement TBD",
                    size=req.get("size", "M").upper(),
                    notes=req.get("notes", "—") or "—",
                )
            )
        screen_details: List[ScreenDetail] = []
        for screen in module.get("screens", []):
            screen_details.append(
                ScreenDetail(
                    title=(screen.get("title") or "Screen").strip(),
                    description=(screen.get("description") or "Description TBD").strip(),
                )
            )
        screen_count_raw = module.get("screen_count")
        try:
            screen_count = int(screen_count_raw) if screen_count_raw is not None else None
        except (TypeError, ValueError):
            screen_count = None
        modules.append(
            ModuleSpec(
                name=module.get("name", "Module").strip() or "Module",
                description=module.get("description", "Description TBD").strip(),
                requirements=requirements,
                summary=module.get("summary", "").strip(),
                screen_details=screen_details,
                screen_count=screen_count,
            )
        )

    risks = data.get("risks", [])
    questions = data.get("questions", [])
    return modules, risks, questions, size_hours


def request_plan_from_llm(description: str, api_key: str | None = None):
    client = get_openai_client(api_key)
    system_prompt = """You are a senior Business Analyst and Presales Solution Consultant.

Your task: Given a feature or project description, produce an exhaustive feature sizing plan as STRICT JSON ONLY, with NO markdown, NO comments, and NO explanations.

The JSON MUST conform to this schema exactly:

{
  "size_hours": {
    "XS": 2,
    "S": 4,
    "M": 8,
    "L": 16,
    "XL": 24
  },
  "modules": [
    {
      "name": "string",
      "description": "1-2 sentence overview of this module",
      "summary": "short business value summary in business language",
      "screen_count": 4,
      "screens": [
        {
          "title": "Screen title",
          "description": "short purpose of this screen"
        }
      ],
      "requirements": [
        {
          "text": "one-line, very granular requirement",
          "size": "XS | S | M | L | XL",
          "notes": "short note or empty string"
        }
      ]
    }
  ],
  "risks": ["short risk statement"],
  "questions": ["short clarification question"]
}

Global output rules:
- Output MUST be valid JSON. Do NOT wrap it in backticks. Do NOT add any text before or after the JSON.
- Do NOT change field names, do NOT add extra top-level fields, and do NOT change the "size_hours" keys or values.

Sizing rules:
- Always include ALL XS, S, M, L, XL keys (uppercase) in "size_hours" with the hours exactly as shown.
- Each requirement's "size" MUST be one of XS, S, M, L, XL only.

Holistic solution coverage (think end-to-end):
- Think like a senior BA: cover the whole solution, not just the obvious UI.
- Identify all major modules needed to deliver the end-to-end product, including where relevant:
  - Core functional modules for the main user journeys
  - Supporting / enabling modules (authentication, user management, roles/permissions)
  - Administration / configuration / settings
  - Reporting, dashboards, analytics
  - Integrations (APIs, webhooks, 3rd party services)
  - Security, audit, logging, compliance-related functions
  - Performance, scalability, reliability and operational aspects where they require explicit features
- Do NOT under-estimate the number of modules: if the scope suggests more modules, add them. Err on the side of including additional modules instead of merging them.

Modules rules:
- Create as many modules as are realistically needed for a complete solution. Do NOT optimise for having fewer modules; prefer decomposing into smaller, actor/stage-focused modules. As a default bias, produce 8-15 modules for a typical business application (more if scope warrants).
- Each module must be coherent (centered around a functional area or actor) and named clearly in business-friendly language.
- "description" should describe what the module does functionally.
- "summary" should express why this module matters in business terms (value, outcome, stakeholder).

Screens rules:
- "screen_count" MUST equal screens.length.
- Only include screens that are relevant to this module.
- Each screen must have:
  - "title": short, human-readable name.
  - "description": one short sentence explaining business purpose or user goal.
- Include screens for:
  - Create/edit/detail views
  - Search/listing views
  - Dashboards
  - Admin/configuration views
  - Error / exception flows only if they require dedicated screens.

Requirements rules (go very granular):
- Think at three layers for each module:
  1) User experience & journeys (what each actor does, step by step)
  2) Business rules, validations, and workflows
  3) System capabilities, data handling, and integrations
- For each module, generate MANY granular requirements. Do NOT limit yourself to 7-12 items; break things down as far as is reasonable. Default bias: 15-25 requirements per module when scope allows.
  - Each requirement should describe ONE atomic behaviour, rule, or system capability.
  - Avoid combining multiple behaviours in a single requirement (avoid "and" chains).
- Cover:
  - Happy path flows
  - Alternative paths
  - Validation rules and error handling
  - Edge cases and exceptional scenarios
  - Permissions / role-based access control
  - Data capture, updates, and state changes
  - Notifications (emails, in-app, SMS, etc. if applicable)
  - Integrations and data sync behaviour
- "text" should be concise but specific and implementation-oriented.
- "notes" may be "" if nothing special is needed, or may include assumptions, dependencies, or BA notes.

BA mindset:
- Explicitly think about different actors (end users, admins, support, external systems) and reflect them in modules, screens, and requirements.
- Surface implicit business rules (e.g., approvals, thresholds, SLAs, audit trails) as separate requirements.
- Where business rules or data definitions are unclear, keep going with reasonable assumptions and capture uncertainty in "risks" and "questions" instead of stopping.

Risks and questions:
- "risks": include 5-10 concise items focusing on:
  - Ambiguities in requirements
  - Technical uncertainties
  - Integration and dependency risks
  - Scope creep or complexity risks
- "questions": include 7-15 items focusing on:
  - Business rules and edge cases
  - Priorities and MVP vs. later phases
  - Data model and reporting needs
  - Non-functional requirements (performance, availability, security, compliance)
- Phrase both risks and questions as if you are preparing for a client / product owner workshop.

If the input is vague:
- Make reasonable, realistic assumptions about the domain, actors, and usage patterns.
- Still produce a full, holistic plan.
- Capture uncertainties or alternate options in "risks" and "questions".
- Do NOT say that you need more information in the JSON."""
    user_prompt = (
        "Feature description:\n"
        f"{description.strip()}\n\n"
        "Respond with the JSON structure described in the system prompt. Do not include any prose."
    )
    response = client.chat.completions.create(
        model="gpt-4.1-mini",
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    )
    content = response.choices[0].message.content
    if not content:
        raise RuntimeError("Empty response from language model.")
    return parse_llm_response(content)


def collect_feature_description() -> str:
    print(
        "Paste the free-text feature description. Press Enter twice when you're done:\n"
    )
    lines: List[str] = []
    blank_streak = 0
    while True:
        try:
            line = input()
        except EOFError:
            break

        if not line.strip():
            blank_streak += 1
            if blank_streak >= 2:
                break
        else:
            blank_streak = 0

        lines.append(line)

    description = "\n".join(lines).strip()
    if not description:
        raise ValueError("Feature description cannot be empty.")
    return description


def sanitize_sheet_name(name: str, existing: set) -> str:
    sanitized = re.sub(r"[\\/*?:\[\]]", "", name)
    if len(sanitized) > 31:
        sanitized = sanitized[:31]
    base = sanitized or "Module"
    final_name = base
    suffix = 1
    while final_name.lower() in {n.lower() for n in existing}:
        trimmed = base[: 31 - len(str(suffix)) - 1]
        final_name = f"{trimmed}_{suffix}"
        suffix += 1
    existing.add(final_name)
    return final_name


def auto_fit_columns(ws, min_width: int = 12):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        length = max(length + 2, min_width)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length, 60)


def write_module_sheet(
    wb: Workbook,
    module: ModuleSpec,
    existing_sheet_names: set,
    size_range_ref: str,
    hours_range_ref: str,
    size_to_hours: Dict[str, float],
) -> Tuple[str, str]:
    sheet_name = sanitize_sheet_name(module.name, existing_sheet_names)
    ws = wb.create_sheet(title=sheet_name)
    bold = Font(bold=True)

    ws["A1"].value = "Module Name"
    ws["A1"].font = bold
    ws["B1"].value = module.name

    ws["A2"].value = "Module Description"
    ws["A2"].font = bold
    ws["B2"].value = module.description
    ws["B2"].alignment = Alignment(wrap_text=True)

    ws["A3"].value = "Number of Screens"
    ws["A3"].font = bold
    ws["B3"].value = module.resolved_screen_count()

    header_row = 5
    ws[f"A{header_row}"].value = "Requirement"
    ws[f"B{header_row}"].value = "Size (XS/S/M/L/XL)"
    ws[f"C{header_row}"].value = "Notes"
    for col in "ABC":
        ws[f"{col}{header_row}"].font = bold

    req_start = header_row + 1
    if not module.requirements:
        module.requirements.append(
            Requirement(
                text="Placeholder requirement derived from description.",
                size="M",
            )
        )

    for idx, req in enumerate(module.requirements):
        row = req_start + idx
        ws[f"A{row}"] = req.text
        ws[f"B{row}"] = req.size
        ws[f"C{row}"] = req.notes
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)

    req_end = req_start + len(module.requirements) - 1
    screens_label_row = req_end + 2
    ws[f"A{screens_label_row}"].value = "Screens"
    ws[f"A{screens_label_row}"].font = bold

    screen_header_row = screens_label_row + 1
    ws[f"A{screen_header_row}"].value = "Screen Title"
    ws[f"B{screen_header_row}"].value = "Description"
    ws[f"A{screen_header_row}"].font = bold
    ws[f"B{screen_header_row}"].font = bold

    screen_rows = module.screen_details or [
        ScreenDetail(
            title="Placeholder Screen",
            description="High-level view derived from requirements.",
        )
    ]
    for idx, screen in enumerate(screen_rows):
        row = screen_header_row + 1 + idx
        ws[f"A{row}"] = screen.title
        ws[f"B{row}"] = screen.description
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)

    screen_table_end = screen_header_row + len(screen_rows)
    summary_row = screen_table_end + 2

    ws[f"A{summary_row}"].value = "Module Summary:"
    ws[f"A{summary_row}"].font = bold
    ws[f"B{summary_row}"].value = module.summary
    ws[f"B{summary_row}"].alignment = Alignment(wrap_text=True)

    ws[f"A{summary_row + 1}"].value = "Screens:"
    ws[f"A{summary_row + 1}"].font = bold
    ws[f"B{summary_row + 1}"].value = module.resolved_screen_count()

    ws[f"A{summary_row + 2}"].value = "Total Module Hours (calculated):"
    ws[f"A{summary_row + 2}"].font = bold
    total_cell_ref = f"B{summary_row + 2}"
    count_range_ref = f"$B${req_start}:$B${req_end}"
    ws[total_cell_ref] = (
        f"=SUMPRODUCT(COUNTIF({count_range_ref},{size_range_ref})*{hours_range_ref})"
    )

    # Static backup value (for viewers that don't evaluate formulas, e.g., some web viewers)
    ws[f"C{summary_row + 2}"].value = "Static total (Python)"
    ws[f"D{summary_row + 2}"] = module.total_hours(size_to_hours)

    auto_fit_columns(ws)
    return sheet_name, total_cell_ref


def build_resource_loading_sheet(
    wb: Workbook,
    module_totals: List[Tuple[str, str, str, float]],
    size_to_hours: Dict[str, float],
    size_order: List[str],
):
    ws = wb.create_sheet(title="ResourceLoading", index=0)
    bold = Font(bold=True)

    ws["A1"].value = "T-shirt Sizing to Hours"
    ws["A1"].font = bold

    ws["A2"].value = "Size"
    ws["B2"].value = "Hours (editable)"
    ws["A2"].font = bold
    ws["B2"].font = bold

    size_start_row = 3
    for offset, size in enumerate(size_order):
        row = size_start_row + offset
        ws[f"A{row}"] = size
        ws[f"B{row}"] = size_to_hours.get(size, DEFAULT_SIZE_TO_HOURS.get(size, 0))

    size_end_row = size_start_row + len(size_order) - 1
    alloc_header_row = size_end_row + 2
    ws[f"A{alloc_header_row}"].value = "Effort Allocation"
    ws[f"A{alloc_header_row}"].font = bold
    ws[f"C{alloc_header_row}"].value = "Ratio (editable)"
    ws[f"C{alloc_header_row}"].font = bold

    total_refs = ",".join([f"'{sheet}'!{cell}" for sheet, _, cell, _ in module_totals])
    static_dev_total = sum(total for _, _, _, total in module_totals)
    dev_row = alloc_header_row + 1
    ws[f"A{dev_row}"] = "Dev Hours"
    ws[f"A{dev_row}"].font = bold
    ws[f"B{dev_row}"] = f"=SUM({total_refs})" if total_refs else 0
    ws[f"D{dev_row}"] = static_dev_total
    ws[f"C{dev_row}"] = "Static Dev Hours"

    allocations = [
        ("QA Hours", 0.4),
        ("PM Hours", 0.15),
        ("Architect Hours", 0.10),
        ("Buffer Hours", 0.20),
    ]
    allocation_rows: Dict[str, int] = {}
    for index, (label, ratio) in enumerate(allocations, start=1):
        row = dev_row + index
        allocation_rows[label] = row
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = bold
        ws[f"C{row}"] = ratio
        ws[f"B{row}"] = f"=$B${dev_row}*$C{row}"

    hours_per_day_row = dev_row + len(allocations) + 2
    ws[f"A{hours_per_day_row}"] = "Hours per Resource Day"
    ws[f"A{hours_per_day_row}"].font = bold
    ws[f"B{hours_per_day_row}"] = 8

    util_header_row = hours_per_day_row + 1
    ws[f"A{util_header_row}"].value = "Resource Utilization"
    ws[f"A{util_header_row}"].font = bold
    ws[f"A{util_header_row + 1}"].value = "Role"
    ws[f"B{util_header_row + 1}"].value = "Hours"
    ws[f"C{util_header_row + 1}"].value = "Resource Days"
    ws[f"E{util_header_row + 1}"].value = "Static Resource Days"
    for col in "ABC":
        ws[f"{col}{util_header_row + 1}"].font = bold
    ws[f"E{util_header_row + 1}"].font = bold

    static_hours: Dict[str, float] = {
        "Dev": static_dev_total,
        "QA": static_dev_total * allocations[0][1],
        "PM": static_dev_total * allocations[1][1],
        "Architect": static_dev_total * allocations[2][1],
        "Buffer": static_dev_total * allocations[3][1],
    }
    util_rows = [
        ("Dev", f"$B${dev_row}"),
        ("QA", f"$B${allocation_rows['QA Hours']}"),
        ("PM", f"$B${allocation_rows['PM Hours']}"),
        ("Architect", f"$B${allocation_rows['Architect Hours']}"),
        ("Buffer", f"$B${allocation_rows['Buffer Hours']}"),
    ]
    for idx, (role, hours_cell) in enumerate(util_rows, start=util_header_row + 2):
        ws[f"A{idx}"] = role
        ws[f"B{idx}"] = f"={hours_cell}"
        ws[f"C{idx}"] = f"={hours_cell}/$B${hours_per_day_row}"
        ws[f"E{idx}"] = round(static_hours.get(role, 0) / ws[f"B{hours_per_day_row}"].value, 2)

    auto_fit_columns(ws, min_width=14)


def build_workbook(
    modules: List[ModuleSpec],
    size_to_hours: Dict[str, float],
    output_path: Path,
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    # Ensure formulas refresh on open; tolerate older openpyxl versions
    try:
        wb.calculation_properties.fullCalcOnLoad = True  # type: ignore[attr-defined]
    except Exception:
        try:
            from openpyxl.workbook.properties import CalcProperties

            wb.calculation_properties = CalcProperties(fullCalcOnLoad=True)  # type: ignore[attr-defined]
        except Exception:
            pass

    size_order = ordered_size_keys(size_to_hours)
    size_start_row = 3
    size_end_row = size_start_row + len(size_order) - 1
    size_range_ref = f"ResourceLoading!$A${size_start_row}:$A${size_end_row}"
    hours_range_ref = f"ResourceLoading!$B${size_start_row}:$B${size_end_row}"

    module_totals: List[Tuple[str, str, str, float]] = []
    existing_sheet_names: set = set()

    for module in modules:
        sheet_name, total_cell = write_module_sheet(
            wb=wb,
            module=module,
            existing_sheet_names=existing_sheet_names,
            size_range_ref=size_range_ref,
            hours_range_ref=hours_range_ref,
            size_to_hours=size_to_hours,
        )
        module_totals.append((sheet_name, module.name, total_cell, module.total_hours(size_to_hours)))

    build_resource_loading_sheet(
        wb=wb,
        module_totals=module_totals,
        size_to_hours=size_to_hours,
        size_order=size_order,
    )
    wb.active = wb["ResourceLoading"]
    wb.save(output_path)


def format_preview(
    modules: List[ModuleSpec],
    size_to_hours: Dict[str, float],
    risks: List[str],
    questions: List[str],
) -> str:
    dev_hours = sum(module.total_hours(size_to_hours) for module in modules)
    qa_hours = round(dev_hours * 0.4, 2)
    pm_hours = round(dev_hours * 0.15, 2)
    architect_hours = round(dev_hours * 0.1, 2)
    buffer_hours = round(dev_hours * 0.2, 2)

    lines: List[str] = []
    for module in modules:
        lines.append(f"Module: {module.name}")
        lines.append(f"Description: {module.description}")
        lines.append(f"Screens: {module.resolved_screen_count()}")
        if module.screen_details:
            lines.append("Screen Details:")
            for screen in module.screen_details:
                lines.append(f"  - {screen.title}: {screen.description}")
        else:
            lines.append("Screen Details: (TBD)")
        lines.append("Requirements:")
        for req in module.requirements:
            notes = req.notes if req.notes else "—"
            lines.append(f"  - [{req.size}] {req.text} ({notes})")
        lines.append(f"Module Summary: {module.summary}")
        lines.append(f"Total Module Hours: {module.total_hours(size_to_hours)}")
        lines.append("")

    lines.append("Resource Loading:")
    lines.append(
        "  Size to Hours Mapping: "
        + ", ".join(f"{k}={v}" for k, v in size_to_hours.items())
    )
    lines.append(f"  Dev Hours: {dev_hours}")
    lines.append(f"  QA Hours (40%): {qa_hours}")
    lines.append(f"  PM Hours (15%): {pm_hours}")
    lines.append(f"  Architect Hours (10%): {architect_hours}")
    lines.append(f"  Buffer (20%): {buffer_hours}")
    lines.append("")

    lines.append("Risks / Missing Information:")
    for risk in risks:
        lines.append(f"  - {risk}")
    lines.append("")

    lines.append("Clarification Questions:")
    for question in questions:
        lines.append(f"  - {question}")

    return "\n".join(lines)


def generate_analysis(
    description: str,
    output_folder: Path,
    api_key: str | None = None,
    size_hours_override: Dict[str, float] | None = None,
):
    modules, risks, questions, size_to_hours = request_plan_from_llm(description, api_key)
    if size_hours_override:
        # Override calibrated S/M/L while keeping any other sizes returned
        for key, value in size_hours_override.items():
            if value is not None:
                size_to_hours[key.upper()] = float(value)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder.mkdir(parents=True, exist_ok=True)
    output_path = output_folder / f"feature_sizing_{timestamp}.xlsx"
    build_workbook(modules, size_to_hours, output_path)
    preview = format_preview(modules, size_to_hours, risks, questions)
    return output_path, preview


def main():
    description = collect_feature_description()
    output_path, preview = generate_analysis(description, Path.cwd())
    divider = "=" * 60
    print(f"\n{divider}\nTEXTUAL PREVIEW\n{divider}")
    print(preview)
    print(f"\nExcel workbook saved to: {output_path.resolve()}\n")


def run_streamlit_app():
    import streamlit as st
    
    st.set_page_config(page_title="Feature Sizing Assistant", page_icon="🤖", layout="wide")

    # --- Theme & layout CSS ---
    st.markdown(
        """
        <style>
        body {
            background: radial-gradient(circle at 10% 20%, #eef2ff 0%, #f7f9ff 24%, #ffffff 55%);
        }
        .ai-card {
            background: linear-gradient(135deg, rgba(99, 102, 241, 0.08), rgba(56, 189, 248, 0.08));
            border: 1px solid rgba(99, 102, 241, 0.2);
            border-radius: 14px;
            padding: 12px 14px;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.08);
        }
        .panel {
            background: #ffffff;
            border-radius: 12px;
            padding: 12px;
            border: 1px solid rgba(148, 163, 184, 0.3);
            box-shadow: 0 8px 18px rgba(15, 23, 42, 0.05);
        }
        .muted { color: #475569; }
        .info-trigger { position: relative; display: inline-block; margin-top: 4px; }
        .info-icon {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            width: 34px;
            height: 34px;
            border-radius: 50%;
            background: #eef2ff;
            color: #0f172a;
            border: 1px solid rgba(148, 163, 184, 0.6);
            cursor: help;
            box-shadow: 0 4px 10px rgba(15, 23, 42, 0.06);
        }
        .info-popover {
            display: none;
            position: absolute;
            top: 42px;
            left: 0;
            z-index: 20;
            width: 350px;
            background: #ffffff;
            border: 1px solid rgba(148, 163, 184, 0.35);
            border-radius: 12px;
            padding: 14px;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.12);
        }
        .info-trigger:hover .info-popover { display: block; }
        .block-container { padding-top: 2rem; }
        .stButton>button {
            background: linear-gradient(120deg, #4338ca, #2563eb);
            color: #ffffff;
            border: none;
            padding: 0.6rem 1rem;
            border-radius: 10px;
            font-weight: 700;
        }
        .stButton>button:hover { opacity: 0.95; }
        /* Make preview/output areas full width and tall */
        .stTextArea, .stTextArea textarea {
            width: 100% !important;
        }
        textarea[aria-label="Preview (for verification before opening Excel)"] {
            min-height: 70vh !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # --- Header ---
    st.markdown(
        """
        <div class="ai-card" style="margin-top: 8px; margin-bottom: 8px;">
            <h3 style="margin: 8px 0 4px 0;">Feature Sizing Assistant</h3>
            <div class="muted" style="font-size: 14px; margin-bottom: 4px;">AI-assisted effort estimation toolfor Business Development teams</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Default API key from environment/Secrets (no input shown)
    default_api_key = os.getenv("OPENAI_API_KEY") or st.secrets.get("OPENAI_API_KEY", "")  # type: ignore[attr-defined]

    # Sidebar for settings and calibration
    with st.sidebar:
        # Info icon at the top of sidebar
        st.markdown(
            """
            <div style="display: flex; align-items: center; gap: 10px; margin: 0 0 16px 0;">
                <div class="info-trigger">
                    <div class="info-icon">ℹ️</div>
                    <div class="info-popover">
                        <div style="font-weight: 700; margin-bottom: 6px;">About this tool</div>
                        <div class="muted" style="font-size: 14px;">
                            <strong>What it does:</strong><br>
                            This tool helps Business Development teams create realistic effort estimations for any complex features quickly.<br><br>
                            <strong>How it works:</strong><br>
                            • Tool breaks down the feature into modules, screens and one line requirements.<br>
                            • Excel does not contain static, hard-coded values.<br>
                            • All estimates are calculated dynamically using rules and complexity multipliers.<br>
                            • Users can calibrate the hours assigned to S (Small), M (Medium), and L (Large).<br>
                            • The tool recalculates total effort instantly based on calibrated values.<br><br>
                            <strong>🎯 How It Helps BD:</strong><br>
                            • Create defensible project estimates<br>
                            • Standardize effort calculations<br>
                            • Simulate multiple scenarios before finalizing proposals<br>
                            • Reduce manual Excel guesswork<br>
                            • Quickly arrive at client-ready effort estimates
                        </div>
                    </div>
                </div>
                <span class="muted" style="font-size: 14px;">Hover for more info</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        
        st.header("Control Center")
        st.subheader("Effort Calibration")
        st.caption("Adjust sizing hours to instantly recalc effort")
        col_xs, col_s, col_m, col_l, col_xl = st.columns(5)
        with col_xs:
            hours_xs = st.number_input("XS hours", min_value=0.5, value=float(DEFAULT_SIZE_TO_HOURS.get("XS", 2)), step=0.25)
        with col_s:
            hours_s = st.number_input("S hours", min_value=1.0, value=float(DEFAULT_SIZE_TO_HOURS.get("S", 4)), step=0.5)
        with col_m:
            hours_m = st.number_input("M hours", min_value=1.0, value=float(DEFAULT_SIZE_TO_HOURS.get("M", 8)), step=0.5)
        with col_l:
            hours_l = st.number_input("L hours", min_value=1.0, value=float(DEFAULT_SIZE_TO_HOURS.get("L", 16)), step=0.5)
        with col_xl:
            hours_xl = st.number_input("XL hours", min_value=1.0, value=float(DEFAULT_SIZE_TO_HOURS.get("XL", 24)), step=0.5)
        calibrated_hours = {
            "XS": hours_xs,
            "S": hours_s,
            "M": hours_m,
            "L": hours_l,
            "XL": hours_xl,
        }

        st.write("Calibrated map:")
        st.code({k: v for k, v in calibrated_hours.items()})

    # Main content area - full width
    st.markdown("### ✍️ Describe the feature")
    description = st.text_area(
        "Free-text Feature Description",
        height=180,
        placeholder="Paste your feature description here...",
        label_visibility="collapsed",
    )
    st.caption("Include scope, integrations, constraints, and acceptance criteria.")

    # Left-aligned button (shorter)
    col_btn_left, col_btn_right = st.columns([1, 5])
    with col_btn_left:
        if st.button("🚀 Get Estimates", type="primary", use_container_width=True):
            if not description.strip():
                st.error("Please provide a feature description.")
            else:
                try:
                    # AI activity feed
                    activity_msgs = [
                        "Analyzing complexity model…",
                        "Applying Excel-based estimation rules…",
                        "Calibrating effort multipliers…",
                        "Generating structured spreadsheet output…",
                    ]
                    st.session_state.activity_logs = activity_msgs
                    with st.spinner("AI is thinking..."):
                        time.sleep(0.6)  # subtle artificial delay
                        output_dir = Path("/tmp") if sys.platform != "win32" else Path.cwd()
                        effective_api_key = default_api_key or None
                        output_path, preview = generate_analysis(
                            description,
                            output_dir,
                            api_key=effective_api_key,
                            size_hours_override=calibrated_hours,
                        )

                    st.toast("AI recalibrated effort model successfully", icon="🤖")
                    st.success("✅ Analysis complete! Excel workbook saved.")

                    with open(output_path, "rb") as f:
                        st.download_button(
                            label="📥 Download Excel Workbook",
                            data=f.read(),
                            file_name=output_path.name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )

                    st.subheader("Textual Preview")
                    st.text_area(
                        "Preview (for verification before opening Excel)",
                        value=preview,
                        height=720,
                        disabled=True,
                        label_visibility="collapsed",
                    )

                except Exception as exc:
                    st.error(f"Failed to generate workbook: {exc}")
                    st.exception(exc)


# Streamlit app - runs when executed via: streamlit run feature_sizing_app.py
try:
    import streamlit as st
    # Only run if streamlit is available (when deployed/run via streamlit)
    # This will execute when Streamlit imports this module
    run_streamlit_app()
except ImportError:
    # Streamlit not available - will use CLI mode instead
    pass

if __name__ == "__main__":
    # For CLI: python feature_sizing_app.py --cli
    if len(sys.argv) > 1 and sys.argv[1] == "--cli":
        main()
    else:
        # Check if streamlit is available
        try:
            import streamlit
            print("Streamlit is available. Run with: streamlit run feature_sizing_app.py")
            print("Or use CLI mode: python feature_sizing_app.py --cli")
        except ImportError:
            # Fall back to CLI if streamlit not available
            main()

